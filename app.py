import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

st.set_page_config(page_title="Limpeza de Excel", layout="centered")
st.title("Limpeza de Excel")

def desmesclar_pagina_1(arquivo_entrada, arquivo_temp):
    wb = load_workbook(arquivo_entrada)
    ws = wb.worksheets[0]  # PRIMEIRA ABA

    for merged in list(ws.merged_cells.ranges):
        valor = ws.cell(
            row=merged.min_row,
            column=merged.min_col
        ).value

        ws.unmerge_cells(str(merged))

        for row in ws.iter_rows(
            min_row=merged.min_row,
            max_row=merged.max_row,
            min_col=merged.min_col,
            max_col=merged.max_col
        ):
            for cell in row:
                cell.value = valor

    wb.save(arquivo_temp)

def tratar_excel(arquivo_entrada, arquivo_saida):
    arquivo_temp = "temp_sem_mescla.xlsx"

    # 1) Remove mesclagem
    desmesclar_pagina_1(arquivo_entrada, arquivo_temp)

    # 2) Lê ignorando linhas 1 a 5 (linha 6 = cabeçalho)
    df = pd.read_excel(
        arquivo_temp,
        sheet_name=0,
        header=5
    )

    # Guarda original
    df_original = df.copy()

    # 3) Coluna I (Compl.lcto) → índice 8
    df.iloc[:, 8] = df.iloc[:, 8].shift(-1)

    # 4) Remove linhas com DT_LANÇTOS vazio (coluna D → índice 3)
    df = df[df.iloc[:, 3].notna()]

    # 5) Gera Excel final
    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ARRUMADO", index=False)
        df_original.to_excel(writer, sheet_name="ORIGINAL", index=False)

    # 6) Limpa temporário
    os.remove(arquivo_temp)

arquivo = st.file_uploader("Envie o arquivo Excel", type=["xlsx"])

if arquivo:
    saida = "arquivo_tratado.xlsx"
    tratar_excel(arquivo, saida)

    with open(saida, "rb") as f:
        st.download_button(
            "Baixar arquivo tratado",
            f,
            file_name="arquivo_tratado.xlsx"
        )

    os.remove(saida)
