import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import os

# ==============================
# CONFIG STREAMLIT
# ==============================
st.set_page_config(
    page_title="Relatório de Despesas",
    layout="wide"
)

st.markdown("## 📊 Relatório de Despesas")
st.markdown(
    "Arquivo tratado automaticamente conforme padrão corporativo."
)

st.divider()

# ==============================
# FUNÇÃO LIMPEZA
# ==============================
def limpar_excel(uploaded_file, arquivo_saida):
    wb = load_workbook(uploaded_file)

    aba_original = wb.worksheets[0]
    copia_original = wb.copy_worksheet(aba_original)
    copia_original.title = "ORIGINAL"

    ws = wb.worksheets[0]

    # 1) Remover mesclagem
    for merged in list(ws.merged_cells.ranges):
        valor = ws.cell(
            row=merged.min_row,
            column=merged.min_col
        ).value
        ws.unmerge_cells(str(merged))
        ws.cell(
            row=merged.min_row,
            column=merged.min_col
        ).value = valor

    # 2) Apagar linhas iniciais
    ws.delete_rows(1, 5)

    # 3) Subir Compl.lcto (coluna I)
    col_compl = 9
    ultima_linha = ws.max_row

    for row in range(3, ultima_linha + 1):
        ws.cell(row=row - 1, column=col_compl).value = (
            ws.cell(row=row, column=col_compl).value
        )

    ws.cell(row=ultima_linha, column=col_compl).value = None

    # 4) Remover "CUSTO C/ TERCEIROS..."
    col_desc_cta = 3
    for row in range(ws.max_row, 1, -1):
        valor = ws.cell(row=row, column=col_desc_cta).value
        if valor and valor.strip() == "CUSTO C/ TERCEIROS PESSOA JURIDI":
            ws.delete_rows(row)

    # 5) Remover linhas sem data
    col_data = 4
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=col_data).value in (None, ""):
            ws.delete_rows(row)

    wb.save(arquivo_saida)

# ==============================
# UPLOAD
# ==============================
arquivo = st.file_uploader(
    "📎 Envie o arquivo Excel de despesas",
    type=["xlsx"]
)

if arquivo:
    data_atual = datetime.now()
    mes_ano = data_atual.strftime("%m/%Y")
    nome_saida = f"RLT_DESPESAS_{data_atual.strftime('%m_%Y')}.xlsx"

    with st.spinner("Processando arquivo..."):
        limpar_excel(arquivo, nome_saida)

    # ==============================
    # LEITURA PARA EXIBIÇÃO
    # ==============================
    df = pd.read_excel(nome_saida)

    st.success("Relatório gerado com sucesso!")

    st.markdown(f"### 🗓️ Período: `{mes_ano}`")
    st.markdown("### 📑 Pré-visualização do relatório")

    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True
    )

    st.divider()

    # ==============================
    # DOWNLOAD
    # ==============================
    with open(nome_saida, "rb") as f:
        st.download_button(
            label="⬇️ Baixar relatório em Excel",
            data=f,
            file_name=nome_saida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    os.remove(nome_saida)
